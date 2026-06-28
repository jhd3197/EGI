"""CSV / Excel (xlsx) bulk import & export for persons (plan-12 §1).

A pragmatic, spreadsheet-friendly companion to the lossless PFIF JSON/XML path
in ``pfif.py``. Operators live in Excel during a crisis, so this module:

- **Exports** a flat, stable column set of person rows to CSV (UTF-8 *with BOM*
  so Excel opens accented Spanish names correctly) or to a real ``.xlsx`` file.
- **Imports** a CSV/xlsx (detected by file extension) with forgiving column
  mapping: a built-in English+Spanish header alias table (``COLUMN_ALIASES``),
  overridable per-call with an explicit ``column_map``. Each row is validated
  (status must be valid if present; name OR cedula required); bad rows are
  collected into an ``errors`` list instead of aborting the whole file.

Imports land as ``source='csv_import'`` with ``reviewed=0`` (unless
``auto_approve``) so bulk uploads await moderation — consistent with
``pfif_import`` and privacy-safe (see ``modules/moderation.UNTRUSTED_SOURCES``).
Exact ``id`` + ``origin_device`` duplicates are skipped, mirroring
``pfif.import_text``.
"""

import csv
import io
import uuid
from typing import List, Optional, Tuple

import db
from models import VALID_STATUSES, now_iso
from modules import provenance

# Stable, flat column set for exports. Snake_case DB column names (CSV/xlsx are
# flat tabular formats, not the /sync JSON envelope, so there is no camelCase
# mapping to honor here).
EXPORT_COLUMNS = [
    "id", "name", "given_name", "family_name", "cedula", "status", "age",
    "sex", "location", "last_seen_date", "contact", "disaster_id", "source",
    "reviewed", "lat", "lon", "created_at", "updated_at",
]

# Canonical person fields an import may set. Anything outside this set is ignored
# (and anything not an actual persons column is dropped before INSERT).
CANONICAL_FIELDS = {
    "id", "name", "given_name", "family_name", "cedula", "status", "age",
    "sex", "gender", "location", "last_seen_date", "contact", "disaster_id",
    "notes", "reporter_name", "origin_device", "lat", "lon",
}

# Friendly blank-template headers handed to operators (auto-managed columns like
# source/reviewed/timestamps are intentionally omitted).
TEMPLATE_HEADERS = [
    "name", "given_name", "family_name", "cedula", "status", "age", "sex",
    "location", "last_seen_date", "contact", "disaster_id", "notes",
]

# Common header aliases (English + Spanish), lowercased, -> canonical field.
# Explicit ``column_map`` entries passed to import_persons override these.
COLUMN_ALIASES = {
    # name
    "nombre": "name", "nombres": "name", "name": "name", "full name": "name",
    "nombre completo": "name",
    "primer nombre": "given_name", "given name": "given_name",
    "given_name": "given_name", "first name": "given_name",
    "apellido": "family_name", "apellidos": "family_name",
    "family name": "family_name", "family_name": "family_name",
    "last name": "family_name", "surname": "family_name",
    # cedula / id document
    "cedula": "cedula", "cédula": "cedula", "ci": "cedula", "dni": "cedula",
    "documento": "cedula", "id document": "cedula",
    # status
    "estado": "status", "status": "status", "estatus": "status",
    # age / sex
    "edad": "age", "age": "age",
    "sexo": "sex", "sex": "sex", "genero": "gender", "género": "gender",
    "gender": "gender",
    # location
    "ubicacion": "location", "ubicación": "location", "location": "location",
    "lugar": "location", "direccion": "location", "dirección": "location",
    "address": "location",
    # date
    "fecha": "last_seen_date", "last_seen_date": "last_seen_date",
    "last seen": "last_seen_date", "fecha visto": "last_seen_date",
    "ultima vez visto": "last_seen_date", "última vez visto": "last_seen_date",
    # contact
    "contacto": "contact", "contact": "contact", "telefono": "contact",
    "teléfono": "contact", "phone": "contact", "celular": "contact",
    # misc
    "notas": "notes", "notes": "notes", "observaciones": "notes",
    "disaster_id": "disaster_id", "desastre": "disaster_id",
    "evento": "disaster_id",
    "reportado por": "reporter_name", "reporter_name": "reporter_name",
    "reporter": "reporter_name",
    "lat": "lat", "latitud": "lat", "latitude": "lat",
    "lon": "lon", "longitud": "lon", "longitude": "lon", "lng": "lon",
    "id": "id",
}

CSV_MEDIA_TYPE = "text/csv; charset=utf-8"
XLSX_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


class OpenpyxlUnavailable(RuntimeError):
    """Raised when an xlsx operation is requested but openpyxl is missing.

    The route layer turns this into a 503 so CSV stays usable even on a minimal
    install without openpyxl.
    """


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------

def _query_persons(status: Optional[str], disaster_id: Optional[str],
                   since: Optional[str]) -> list:
    sql = f"SELECT {', '.join(EXPORT_COLUMNS)} FROM persons WHERE 1=1"
    params: list = []
    if status:
        sql += " AND status = ?"
        params.append(status)
    if disaster_id:
        sql += " AND disaster_id = ?"
        params.append(disaster_id)
    if since:
        sql += " AND updated_at > ?"
        params.append(since)
    sql += " ORDER BY updated_at ASC"
    with db.get_db() as conn:
        return [db.row_to_dict(r) for r in conn.execute(sql, params).fetchall()]


def _persons_to_csv(rows: list) -> bytes:
    """Serialize person rows to CSV bytes with a UTF-8 BOM (Excel-friendly)."""
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=EXPORT_COLUMNS, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow({c: ("" if row.get(c) is None else row.get(c)) for c in EXPORT_COLUMNS})
    # utf-8-sig prepends the BOM so Excel renders accented names correctly.
    return buf.getvalue().encode("utf-8-sig")


def _persons_to_xlsx(rows: list) -> bytes:
    try:
        from openpyxl import Workbook
    except ImportError as e:  # pragma: no cover - exercised via the 503 path
        raise OpenpyxlUnavailable(
            "openpyxl is required for xlsx export/import; install it or use CSV."
        ) from e
    wb = Workbook()
    ws = wb.active
    ws.title = "persons"
    ws.append(EXPORT_COLUMNS)
    for row in rows:
        ws.append([row.get(c) for c in EXPORT_COLUMNS])
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def export_persons(fmt: str = "csv", status: Optional[str] = None,
                   disaster_id: Optional[str] = None,
                   since: Optional[str] = None) -> Tuple[bytes, str, str]:
    """Export persons as CSV or xlsx.

    Returns ``(content_bytes, filename, media_type)``. ``fmt`` is ``csv`` or
    ``xlsx``; raises ``OpenpyxlUnavailable`` if xlsx is requested without
    openpyxl, and ``ValueError`` for an unknown format.
    """
    db.init_db()
    rows = _query_persons(status, disaster_id, since)
    if fmt == "xlsx":
        return _persons_to_xlsx(rows), "egi-persons.xlsx", XLSX_MEDIA_TYPE
    if fmt == "csv":
        return _persons_to_csv(rows), "egi-persons.csv", CSV_MEDIA_TYPE
    raise ValueError(f"Unsupported export format: {fmt!r}")


def template_csv() -> bytes:
    """A blank CSV template (canonical headers only) for operators to fill in."""
    buf = io.StringIO()
    csv.writer(buf).writerow(TEMPLATE_HEADERS)
    return buf.getvalue().encode("utf-8-sig")


# ---------------------------------------------------------------------------
# Import
# ---------------------------------------------------------------------------

def _parse_csv(file_bytes: bytes) -> Tuple[List[str], List[dict]]:
    # utf-8-sig transparently strips a BOM written by Excel/our own export.
    text = file_bytes.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    headers = reader.fieldnames or []
    return list(headers), [dict(row) for row in reader]


def _parse_xlsx(file_bytes: bytes) -> Tuple[List[str], List[dict]]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:  # pragma: no cover - exercised via the 503 path
        raise OpenpyxlUnavailable(
            "openpyxl is required for xlsx export/import; install it or use CSV."
        ) from e
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], []
    headers = [("" if h is None else str(h)) for h in header_row]
    records = []
    for raw in rows_iter:
        if raw is None or all(c is None for c in raw):
            continue
        records.append({headers[i]: raw[i] for i in range(min(len(headers), len(raw)))})
    return headers, records


def _resolve_field(header: str, column_map: Optional[dict]) -> Optional[str]:
    """Map a source header to a canonical person field, or None to ignore it."""
    if header is None:
        return None
    key = str(header).strip()
    if column_map and key in column_map:
        return column_map[key]
    low = key.lower()
    if low in COLUMN_ALIASES:
        return COLUMN_ALIASES[low]
    if low in CANONICAL_FIELDS:
        return low
    return None


def _map_row(raw_row: dict, column_map: Optional[dict]) -> dict:
    mapped: dict = {}
    for header, value in raw_row.items():
        field = _resolve_field(header, column_map)
        if not field:
            continue
        if isinstance(value, str):
            value = value.strip()
        if value in (None, ""):
            continue
        mapped[field] = value
    # Compose a display name from given/family parts when no explicit name given.
    if not mapped.get("name"):
        parts = [mapped.get("given_name"), mapped.get("family_name")]
        composed = " ".join(str(p) for p in parts if p).strip()
        if composed:
            mapped["name"] = composed
    return mapped


def import_persons(file_bytes: bytes, filename: str, column_map: Optional[dict] = None,
                   auto_approve: bool = False, dry_run: bool = False,
                   disaster_id: Optional[str] = None,
                   uploaded_by: str = "csv-import") -> dict:
    """Parse + import a CSV/xlsx file of persons.

    Returns ``{"saved", "skipped", "errors", "total", "batch_id"}``.
    Row numbers in ``errors`` are 1-based spreadsheet lines (header = line 1).
    """
    db.init_db()
    name = (filename or "").lower()
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        extraction_method = "openpyxl"
        _, raw_rows = _parse_xlsx(file_bytes)
    else:
        extraction_method = "stdlib-csv"
        _, raw_rows = _parse_csv(file_bytes)

    reviewed_val = 1 if auto_approve else 0
    now = now_iso()
    saved = skipped = 0
    errors: list = []
    total = len(raw_rows)
    batch_id: Optional[str] = None

    with db.get_db() as conn:
        cur = conn.cursor()
        person_cols = {r[1] for r in cur.execute("PRAGMA table_info(persons)").fetchall()}

        # Create the provenance batch before processing rows. Even on dry_run we
        # create it so the caller can inspect what would have been imported.
        batch_id = provenance.create_import_batch(
            conn,
            file_bytes=file_bytes,
            source_type="csv_import",
            extraction_method=extraction_method,
            original_filename=filename,
            media_type="text/csv" if extraction_method == "stdlib-csv" else XLSX_MEDIA_TYPE,
            disaster_id=disaster_id,
            uploaded_by=uploaded_by,
        )

        for i, raw in enumerate(raw_rows):
            line = i + 2  # +1 for 0-index, +1 for the header line
            mapped = _map_row(raw, column_map)

            # Validation: status (if present) must be valid; need name OR cedula.
            status = mapped.get("status")
            if status is not None and status not in VALID_STATUSES:
                errors.append({"row": line, "error": f"invalid status '{status}'"})
                continue
            if not mapped.get("name") and not mapped.get("cedula"):
                errors.append({"row": line, "error": "row needs a name or cedula"})
                continue

            pid = mapped.get("id") or f"egi-csv-{uuid.uuid4().hex[:10]}"
            origin = mapped.get("origin_device")
            # Skip exact id + origin_device duplicates (mirrors pfif.import_text).
            if mapped.get("id"):
                dup = cur.execute(
                    "SELECT 1 FROM persons WHERE id = ? AND IFNULL(origin_device,'') = IFNULL(?,'')",
                    (pid, origin),
                ).fetchone()
                if dup:
                    skipped += 1
                    continue

            # Coerce numeric fields; a bad value is an error row, not a crash.
            try:
                rec = _build_record(mapped, person_cols, pid, reviewed_val, now, batch_id, filename)
            except ValueError as e:
                errors.append({"row": line, "error": str(e)})
                continue

            if not dry_run:
                cols = ", ".join(rec.keys())
                ph = ", ".join(f":{k}" for k in rec.keys())
                cur.execute(f"INSERT INTO persons ({cols}) VALUES ({ph})", rec)
            saved += 1

        # Finalize batch status/count.
        provenance.finalize_batch(conn, batch_id, saved, errors)

        if not dry_run:
            conn.commit()

    return {
        "saved": saved,
        "skipped": skipped,
        "errors": errors,
        "total": total,
        "batch_id": batch_id,
    }


def _build_record(mapped: dict, person_cols: set, pid: str, reviewed_val: int,
                  now: str, batch_id: Optional[str] = None,
                  filename: str = "CSV/Excel file") -> dict:
    """Build an INSERT-ready person dict from a mapped row (raises ValueError)."""
    rec: dict = {}
    for key, value in mapped.items():
        if key not in person_cols:
            continue
        if key == "age":
            try:
                value = int(value)
            except (ValueError, TypeError):
                raise ValueError(f"invalid age '{value}'")
        elif key in ("lat", "lon"):
            try:
                value = float(value)
            except (ValueError, TypeError):
                raise ValueError(f"invalid {key} '{value}'")
        rec[key] = value
    rec["id"] = pid
    rec["source"] = "csv_import"
    rec["reviewed"] = reviewed_val
    rec["provenance"] = f"Imported from '{filename}' via CSV/Excel on {now}"
    rec["import_batch_id"] = batch_id
    rec["created_at"] = now
    rec["updated_at"] = now
    return rec


def validation_report_csv(errors: list) -> bytes:
    """Render an import errors list to a downloadable CSV (UTF-8 BOM)."""
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=["row", "error"], extrasaction="ignore")
    writer.writeheader()
    for e in errors:
        writer.writerow({"row": e.get("row"), "error": e.get("error")})
    return buf.getvalue().encode("utf-8-sig")
