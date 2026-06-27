# CSV/Excel import & export (plan-12 §1). Exercises modules/exchange.py logic
# plus the HTTP routes via TestClient. TEST DATA — NOT REAL.
import csv
import io

import db
from modules import exchange
from seed import seed as seedmod


def _count(table):
    with db.get_db() as conn:
        return conn.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]


# ── Export ───────────────────────────────────────────────────────────────────

def test_csv_export_roundtrips_header_and_rows(temp_db):
    seedmod.seed_database(person_count=6, report_count=0)
    content, filename, media_type = exchange.export_persons("csv")
    assert filename.endswith(".csv")
    assert "text/csv" in media_type
    # utf-8-sig BOM so Excel renders accented names correctly.
    assert content.startswith(b"\xef\xbb\xbf")
    text = content.decode("utf-8-sig")
    rows = list(csv.DictReader(io.StringIO(text)))
    assert rows[0].keys() >= {"id", "name", "status", "cedula"}
    assert len(rows) == 6


def test_xlsx_export_is_openpyxl_readable(temp_db):
    seedmod.seed_database(person_count=4, report_count=0)
    content, filename, media_type = exchange.export_persons("xlsx")
    assert filename.endswith(".xlsx")
    assert isinstance(content, bytes) and content[:2] == b"PK"  # zip/xlsx magic
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0][:2] == ("id", "name")
    assert len(rows) == 1 + 4  # header + persons


def test_export_filters_by_status(temp_db):
    seedmod.seed_database(person_count=10, report_count=0)
    content, _, _ = exchange.export_persons("csv", status="missing")
    rows = list(csv.DictReader(io.StringIO(content.decode("utf-8-sig"))))
    assert all(r["status"] == "missing" for r in rows)


# ── Import ───────────────────────────────────────────────────────────────────

def _csv_bytes(header, *rows):
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(header)
    for r in rows:
        w.writerow(r)
    return buf.getvalue().encode("utf-8")


def test_import_maps_columns_and_saves(temp_db):
    data = _csv_bytes(
        ["name", "status", "cedula", "age", "location"],
        ["Ana Pérez", "missing", "V-12345678", "30", "Caracas"],
        ["Luis Gómez", "safe", "V-87654321", "45", "Valencia"],
    )
    result = exchange.import_persons(data, "people.csv")
    assert result["total"] == 2
    assert result["saved"] == 2
    assert result["errors"] == []
    assert _count("persons") == 2
    with db.get_db() as conn:
        row = conn.execute("SELECT source, reviewed, age FROM persons WHERE name = ?",
                           ("Ana Pérez",)).fetchone()
    assert row["source"] == "csv_import"
    assert row["reviewed"] == 0  # awaits moderation by default
    assert row["age"] == 30


def test_import_invalid_status_goes_to_errors(temp_db):
    data = _csv_bytes(
        ["name", "status"],
        ["Valid Person", "missing"],
        ["Bad Person", "nonsense"],
    )
    result = exchange.import_persons(data, "people.csv")
    assert result["saved"] == 1
    assert len(result["errors"]) == 1
    assert result["errors"][0]["row"] == 3  # header is line 1
    assert "invalid status" in result["errors"][0]["error"]


def test_import_requires_name_or_cedula(temp_db):
    data = _csv_bytes(["name", "status", "location"], ["", "missing", "Caracas"])
    result = exchange.import_persons(data, "people.csv")
    assert result["saved"] == 0
    assert len(result["errors"]) == 1
    assert "name or cedula" in result["errors"][0]["error"]


def test_import_skips_duplicate_id(temp_db):
    data = _csv_bytes(["id", "name", "status"], ["egi-dup-1", "Once", "safe"])
    first = exchange.import_persons(data, "p.csv")
    assert first["saved"] == 1
    second = exchange.import_persons(data, "p.csv")
    assert second["saved"] == 0
    assert second["skipped"] == 1
    assert _count("persons") == 1


def test_import_auto_approve_sets_reviewed(temp_db):
    data = _csv_bytes(["name", "status"], ["Trusted", "safe"])
    exchange.import_persons(data, "p.csv", auto_approve=True)
    with db.get_db() as conn:
        reviewed = conn.execute("SELECT reviewed FROM persons").fetchone()[0]
    assert reviewed == 1


def test_import_spanish_aliases(temp_db):
    # Spanish headers map via COLUMN_ALIASES with no explicit column_map.
    data = _csv_bytes(
        ["nombre", "estado", "edad", "ubicacion"],
        ["Carlos Ruiz", "found", "22", "Maracaibo"],
    )
    result = exchange.import_persons(data, "personas.csv")
    assert result["saved"] == 1
    with db.get_db() as conn:
        row = conn.execute("SELECT name, status, age, location FROM persons").fetchone()
    assert row["name"] == "Carlos Ruiz"
    assert row["status"] == "found"
    assert row["age"] == 22
    assert row["location"] == "Maracaibo"


def test_import_dry_run_writes_nothing(temp_db):
    data = _csv_bytes(["name", "status"], ["Ghost", "missing"])
    result = exchange.import_persons(data, "p.csv", dry_run=True)
    assert result["saved"] == 1
    assert _count("persons") == 0


def test_explicit_column_map_overrides_aliases(temp_db):
    data = _csv_bytes(["full_legal_name", "estado"], ["Maria Silva", "care"])
    result = exchange.import_persons(
        data, "p.csv", column_map={"full_legal_name": "name"}
    )
    assert result["saved"] == 1
    with db.get_db() as conn:
        row = conn.execute("SELECT name, status FROM persons").fetchone()
    assert row["name"] == "Maria Silva"
    assert row["status"] == "care"


# ── HTTP routes ──────────────────────────────────────────────────────────────

def test_http_export_csv(client):
    seedmod.seed_database(person_count=3, report_count=0)
    resp = client.get("/export/persons.csv")
    assert resp.status_code == 200
    assert "attachment" in resp.headers["content-disposition"]
    rows = list(csv.DictReader(io.StringIO(resp.content.decode("utf-8-sig"))))
    assert len(rows) == 3


def test_http_template_csv(client):
    resp = client.get("/import/persons/template.csv")
    assert resp.status_code == 200
    header = resp.content.decode("utf-8-sig").splitlines()[0]
    assert "name" in header and "cedula" in header


def test_http_import_persons(client):
    data = _csv_bytes(["nombre", "estado"], ["Pedro Loera", "missing"])
    resp = client.post(
        "/import/persons",
        files={"file": ("personas.csv", data, "text/csv")},
        data={"auto_approve": "true"},
    )
    assert resp.status_code == 200
    body = resp.json()
    assert body["saved"] == 1
    assert _count("persons") == 1
